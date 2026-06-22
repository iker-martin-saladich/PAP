#!/usr/bin/env python3
"""
actualitzar.py
Llegeix els Excel PLM i incrusta les dades directament als HTML.
Funciona tant localment com a GitHub Actions.

Ús:  python actualitzar.py
"""

from pathlib import Path
from openpyxl import load_workbook
import re

BASE = Path(__file__).parent

# ── Helpers de lectura ────────────────────────────────────────────────────────

def cel(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v

def es_num(v):
    try:
        int(str(v).strip())
        return True
    except:
        return False

def files_dades(ws, fila_ini=5):
    return [r for r in range(fila_ini, ws.max_row + 1)
            if es_num(ws.cell(r, 1).value)]

def pct_num(v):
    try:
        return float(v)
    except:
        return 0.0

def progress_global(ws_fases, c_pct):
    vals = [pct_num(cel(ws_fases, r, c_pct)) for r in files_dades(ws_fases)]
    return round(sum(vals) / len(vals), 1) if vals else 0

# ── Generadors HTML ───────────────────────────────────────────────────────────

ESTAT_COLOR = {
    "Completat": ("#4A7C59", "#E8F2EB"),
    "En progrés": ("#8B5E3C", "#F0E6D6"),
    "Pendent":   ("#8C7B68", "#F0EEEB"),
}
PRIOR_COLOR = {
    "Alta":  ("#8B3A3A", "#FDEAEA"),
    "Mitja": ("#A07820", "#FEF6E0"),
    "Baixa": ("#4A7C59", "#E8F2EB"),
}
RISC_COLOR = {
    "Alta":  ("#8B3A3A", "#FDEAEA"),
    "Mitja": ("#A07820", "#FEF6E0"),
    "Baixa": ("#4A7C59", "#E8F2EB"),
}
VOCAB_COLOR = {
    "Après":    ("#4A7C59", "#E8F2EB"),
    "En repàs": ("#A07820", "#FEF6E0"),
    "Pendent":  ("#8C7B68", "#F0EEEB"),
}

def badge(text, fg, bg):
    return (f'<span style="font-size:0.7rem;font-weight:500;padding:0.2rem 0.65rem;'
            f'border-radius:2rem;background:{bg};color:{fg};white-space:nowrap">{text}</span>')

def barra(pct, color="#8B5E3C"):
    p = min(100, max(0, pct_num(pct)))
    return (f'<div style="height:5px;background:#E2D9CE;border-radius:3px;overflow:hidden;margin-top:0.4rem">'
            f'<div style="width:{p:.0f}%;height:100%;background:{color};border-radius:3px"></div></div>'
            f'<div style="font-size:0.72rem;color:#8C7B68;text-align:right;margin-top:2px">{p:.0f}%</div>')

def html_fases(ws, c_titol, c_desc, c_pct, c_estat, color_accent="#8B5E3C"):
    parts = []
    for r in files_dades(ws):
        titol = cel(ws, r, c_titol)
        desc  = cel(ws, r, c_desc)
        pct   = pct_num(cel(ws, r, c_pct))
        estat = str(cel(ws, r, c_estat))
        fg, bg = ESTAT_COLOR.get(estat, ("#8C7B68", "#F0EEEB"))
        icon  = {"Completat": "✅", "En progrés": "🔧", "Pendent": "⏳"}.get(estat, "⏳")
        parts.append(
            f'<div style="display:flex;align-items:center;gap:1rem;padding:0.9rem 1.1rem;'
            f'background:#FEFCF9;border:1px solid #E2D9CE;border-radius:0.75rem;margin-bottom:0.6rem">'
            f'<span style="font-size:1rem">{icon}</span>'
            f'<div style="flex:1">'
            f'<strong style="font-size:0.88rem;display:block">{titol}</strong>'
            f'<span style="font-size:0.76rem;color:#8C7B68">{desc}</span>'
            f'</div>'
            f'{badge(estat, fg, bg)}'
            f'<div style="width:80px">{barra(pct, fg)}</div>'
            f'</div>'
        )
    return "\n".join(parts) if parts else "<p style='color:#8C7B68'>Sense dades.</p>"

def html_tasques(ws, c_titol, c_fase, c_prior, c_estat, c_pct, c_notes):
    parts = []
    for r in files_dades(ws):
        titol = str(cel(ws, r, c_titol))
        fase  = str(cel(ws, r, c_fase))
        prior = str(cel(ws, r, c_prior))
        estat = str(cel(ws, r, c_estat))
        notes = str(cel(ws, r, c_notes))
        done  = estat == "Completat"
        fg_p, bg_p = PRIOR_COLOR.get(prior, ("#8C7B68", "#F0EEEB"))
        td = 'text-decoration:line-through;color:#8C7B68' if done else 'color:#1A1208'
        notes_html = (f'<span style="font-size:0.72rem;color:#8C7B68">{notes}</span>'
                      if notes else "")
        parts.append(
            f'<div style="display:flex;gap:0.75rem;padding:0.6rem 0.4rem;border-bottom:1px solid #E2D9CE">'
            f'<span style="font-size:0.85rem;margin-top:1px">{"✅" if done else "⬜"}</span>'
            f'<div style="flex:1">'
            f'<div style="font-size:0.86rem;{td}">{titol}</div>'
            f'<div style="display:flex;gap:0.6rem;margin-top:0.15rem;flex-wrap:wrap">'
            f'<span style="font-size:0.72rem;color:#8C7B68">{fase}</span>'
            f'{badge(prior, fg_p, bg_p)}'
            f'{notes_html}'
            f'</div></div></div>'
        )
    return "\n".join(parts) if parts else "<p style='color:#8C7B68'>Sense dades.</p>"

def html_riscos(ws, c_risc, c_prob, c_imp, c_mitig, c_estat):
    parts = []
    for r in files_dades(ws):
        risc  = str(cel(ws, r, c_risc))
        prob  = str(cel(ws, r, c_prob))
        mitig = str(cel(ws, r, c_mitig))
        fg, bg = RISC_COLOR.get(prob, ("#8C7B68", "#F0EEEB"))
        parts.append(
            f'<div style="display:flex;gap:0.75rem;padding:0.65rem 0;border-bottom:1px solid #E2D9CE;font-size:0.85rem">'
            f'{badge(prob, fg, bg)}'
            f'<div>'
            f'<div style="color:#4A3F2F">{risc}</div>'
            f'<div style="font-size:0.75rem;color:#8C7B68">Mitigació: {mitig}</div>'
            f'</div></div>'
        )
    return "\n".join(parts) if parts else "<p style='color:#8C7B68'>Sense dades.</p>"

def html_pressupost(ws, c_conc, c_cat, c_est, c_notes):
    parts = []
    total_est = 0
    for r in files_dades(ws):
        conc = str(cel(ws, r, c_conc))
        cat  = str(cel(ws, r, c_cat)) if c_cat else ""
        est  = cel(ws, r, c_est)
        try:
            v = float(est)
            total_est += v
            est_txt = f"{v:.0f}€"
        except:
            est_txt = "—"
        parts.append(
            f'<div style="display:flex;justify-content:space-between;align-items:center;'
            f'padding:0.5rem 0;border-bottom:1px solid #E2D9CE;font-size:0.85rem">'
            f'<span style="color:#4A3F2F">{conc}'
            f'<span style="font-size:0.68rem;color:#8C7B68;background:#FAF7F2;padding:0.1rem 0.5rem;'
            f'border-radius:2rem;margin-left:0.5rem">{cat}</span></span>'
            f'<span style="font-family:\'Playfair Display\',serif;color:#5C4A1A">{est_txt}</span>'
            f'</div>'
        )
    parts.append(
        f'<div style="display:flex;justify-content:space-between;padding:0.65rem 0;'
        f'border-top:2px solid #E2D9CE;margin-top:0.25rem;font-weight:500;font-size:0.9rem">'
        f'<span>Total estimat</span>'
        f'<span style="font-family:\'Playfair Display\',serif;color:#5C4A1A;font-size:1.05rem">'
        f'{total_est:.0f}€</span></div>'
    )
    return "\n".join(parts)

def html_bibliografia(ws):
    parts = []
    for r in files_dades(ws):
        any_   = cel(ws, r, 2)
        ref    = str(cel(ws, r, 3))
        tema   = str(cel(ws, r, 4))
        llegit = str(cel(ws, r, 5))
        citat  = str(cel(ws, r, 6))
        if not ref or "Afegeix" in ref:
            continue
        llegit_html = ('<span style="font-size:0.68rem;color:#4A7C59">· Llegit</span>'
                       if llegit == "Sí" else
                       '<span style="font-size:0.68rem;color:#8C7B68">· Pendent</span>')
        citat_html  = ('<span style="font-size:0.68rem;color:#3A5C8B">· Citat</span>'
                       if citat == "Sí" else "")
        parts.append(
            f'<div style="display:flex;gap:0.75rem;padding:0.65rem 0;border-bottom:1px solid #E2D9CE">'
            f'<div style="width:6px;height:6px;min-width:6px;border-radius:50%;'
            f'background:#3A5C8B;margin-top:0.55rem"></div>'
            f'<div>'
            f'<div style="font-size:0.85rem;color:#4A3F2F;line-height:1.6">'
            f'{"(" + str(any_) + ") " if any_ else ""}{ref}</div>'
            f'<div style="font-size:0.72rem;color:#8C7B68;margin-top:0.1rem">'
            f'{tema} {llegit_html} {citat_html}</div>'
            f'</div></div>'
        )
    return "\n".join(parts) if parts else "<p style='color:#8C7B68'>Sense referències.</p>"

def html_gramatica(ws):
    parts = []
    for r in files_dades(ws):
        tema  = str(cel(ws, r, 2))
        desc  = str(cel(ws, r, 3))
        fase  = str(cel(ws, r, 4))
        estat = str(cel(ws, r, 5))
        done  = estat == "Completat"
        fg, bg = ESTAT_COLOR.get(estat, ("#8C7B68", "#F0EEEB"))
        td = 'text-decoration:line-through;color:#8C7B68' if done else 'color:#1A1208'
        parts.append(
            f'<div style="display:flex;gap:0.75rem;padding:0.6rem 0.4rem;border-bottom:1px solid #E2D9CE">'
            f'<span style="font-size:0.85rem;margin-top:1px">{"✅" if done else "⬜"}</span>'
            f'<div style="flex:1">'
            f'<div style="font-size:0.86rem;{td}">{tema}</div>'
            f'<div style="display:flex;gap:0.6rem;margin-top:0.15rem;flex-wrap:wrap">'
            f'<span style="font-size:0.72rem;color:#8C7B68;font-style:italic">{desc}</span>'
            f'{badge(fase, "#1B3A5C", "#E6EEF6")}'
            f'</div></div>'
            f'{badge(estat, fg, bg)}'
            f'</div>'
        )
    return "\n".join(parts) if parts else "<p style='color:#8C7B68'>Sense dades.</p>"

def html_vocabulari(ws):
    parts = []
    for r in files_dades(ws):
        fr    = str(cel(ws, r, 2))
        cat   = str(cel(ws, r, 3))
        ex    = str(cel(ws, r, 4))
        categ = str(cel(ws, r, 5))
        estat = str(cel(ws, r, 6))
        if not fr or not fr.strip():
            continue
        fg, bg = VOCAB_COLOR.get(estat, ("#8C7B68", "#F0EEEB"))
        parts.append(
            f'<div style="display:flex;align-items:center;gap:0.75rem;padding:0.55rem 0.4rem;'
            f'border-bottom:1px solid #E2D9CE;font-size:0.85rem">'
            f'<div style="flex:1">'
            f'<span style="font-weight:500;color:#1B3A5C">{fr}</span>'
            f'<span style="color:#8C7B68;margin:0 0.4rem">→</span>'
            f'<span style="color:#4A3F2F">{cat}</span>'
            f'<div style="font-size:0.72rem;color:#8C7B68;font-style:italic;margin-top:1px">{ex}</div>'
            f'</div>'
            f'{badge(categ, "#4A3F2F", "#F0EEEB")}'
            f'{badge(estat, fg, bg)}'
            f'</div>'
        )
    return "\n".join(parts) if parts else "<p style='color:#8C7B68'>Sense dades.</p>"

def html_registre(ws, max_files=10):
    parts = []
    totes = files_dades(ws)
    # Mostrar les últimes max_files sessions no buides
    sessions = [r for r in totes if cel(ws, r, 2) and str(cel(ws, r, 2)).strip()]
    sessions = sessions[-max_files:]
    for r in reversed(sessions):
        data  = str(cel(ws, r, 2))
        dia   = str(cel(ws, r, 3))
        hores = cel(ws, r, 4)
        acts  = str(cel(ws, r, 5))
        vocab = str(cel(ws, r, 6))
        obs   = str(cel(ws, r, 7))
        try:
            h_txt = f"{float(hores):.1f}h"
        except:
            h_txt = "—"
        parts.append(
            f'<div style="padding:0.9rem 0;border-bottom:1px solid #E2D9CE">'
            f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:0.4rem">'
            f'<div>'
            f'<span style="font-weight:500;font-size:0.88rem;color:#1A1208">{data}</span>'
            f'<span style="font-size:0.75rem;color:#8C7B68;margin-left:0.5rem">{dia}</span>'
            f'</div>'
            f'<span style="font-family:\'Playfair Display\',serif;color:#1B3A5C;font-size:1rem">{h_txt}</span>'
            f'</div>'
            f'<div style="font-size:0.82rem;color:#4A3F2F">{acts}</div>'
            f'{"<div style=\\'font-size:0.78rem;color:#4A7C59;margin-top:0.2rem\\'>📚 " + vocab + "</div>" if vocab else ""}'
            f'{"<div style=\\'font-size:0.75rem;color:#8C7B68;font-style:italic;margin-top:0.2rem\\'>" + obs + "</div>" if obs else ""}'
            f'</div>'
        )
    return "\n".join(parts) if parts else "<p style='color:#8C7B68'>Encara no hi ha sessions registrades.</p>"

# ── Injecció HTML ─────────────────────────────────────────────────────────────

def injectar(html_path, seccions):
    text = html_path.read_text(encoding="utf-8")
    for clau, contingut in seccions.items():
        pattern = rf'<!--\s*PLM:{clau}\s*-->.*?<!--\s*/PLM:{clau}\s*-->'
        replacement = f'<!-- PLM:{clau} -->\n{contingut}\n<!-- /PLM:{clau} -->'
        text, n = re.subn(pattern, replacement, text, flags=re.DOTALL)
        if n == 0:
            print(f"  ⚠  Marca PLM:{clau} no trobada a {html_path.name}")
    html_path.write_text(text, encoding="utf-8")

# ── Processadors per projecte ─────────────────────────────────────────────────

def processar_web():
    xlsx = BASE / "web_personal" / "PLM_Web_Personal.xlsx"
    html = BASE / "web_personal" / "plm_web.html"
    if not xlsx.exists(): print("  ⚠  Excel no trobat"); return
    wb = load_workbook(xlsx, data_only=True)
    ws_f = wb["Fases"]; ws_t = wb["Tasques"]
    ws_r = wb["Riscos"]; ws_p = wb["Pressupost"]
    pg = progress_global(ws_f, 6)
    fases_done = sum(1 for r in files_dades(ws_f) if cel(ws_f,r,8) == "Completat")
    task_done  = sum(1 for r in files_dades(ws_t) if cel(ws_t,r,5) == "Completat")
    task_total = len(files_dades(ws_t))
    riscos_oberts = sum(1 for r in files_dades(ws_r)
                        if cel(ws_r,r,7) not in ("Tancat","Acceptat"))
    seccions = {
        "PROGRESS_PCT": f'{pg:.0f}%',
        "PROGRESS_BAR": f'{pg:.0f}',
        "H_PROGRESS":   f'{pg:.0f}%',
        "H_FASES":      f'{fases_done}/{len(files_dades(ws_f))}',
        "H_TASQUES":    f'{task_done}/{task_total}',
        "H_RISCOS":     str(riscos_oberts),
        "FASES":        html_fases(ws_f, 2, 3, 6, 8),
        "TASQUES":      html_tasques(ws_t, 2, 3, 4, 5, 6, 7),
        "RISCOS":       html_riscos(ws_r, 2, 3, 4, 5, 7),
        "PRESSUPOST":   html_pressupost(ws_p, 2, 3, 4, 7),
    }
    injectar(html, seccions)
    print(f"  ✓  plm_web.html  →  {pg:.0f}% | tasques {task_done}/{task_total}")

def processar_tfm():
    xlsx = BASE / "tfm_quantica" / "PLM_TFM_Quantica.xlsx"
    html = BASE / "tfm_quantica" / "plm_tesi.html"
    if not xlsx.exists(): print("  ⚠  Excel no trobat"); return
    wb = load_workbook(xlsx, data_only=True)
    ws_c = wb["Capítols"]; ws_t = wb["Tasques"]
    ws_r = wb["Riscos"];   ws_b = wb["Bibliografia"]
    pg = progress_global(ws_c, 6)
    caps_done  = sum(1 for r in files_dades(ws_c) if cel(ws_c,r,8) == "Completat")
    task_done  = sum(1 for r in files_dades(ws_t) if cel(ws_t,r,5) == "Completat")
    task_total = len(files_dades(ws_t))
    refs_total = sum(1 for r in files_dades(ws_b)
                     if cel(ws_b,r,3) and "Afegeix" not in str(cel(ws_b,r,3)))
    seccions = {
        "PROGRESS_PCT": f'{pg:.0f}%',
        "PROGRESS_BAR": f'{pg:.0f}',
        "H_PROGRESS":   f'{pg:.0f}%',
        "H_CAPS":       f'{caps_done}/{len(files_dades(ws_c))}',
        "H_TASQUES":    f'{task_done}/{task_total}',
        "H_REFS":       str(refs_total),
        "FASES":        html_fases(ws_c, 2, 3, 6, 8, "#3A5C8B"),
        "TASQUES":      html_tasques(ws_t, 2, 3, 4, 5, 6, 7),
        "RISCOS":       html_riscos(ws_r, 2, 3, 4, 5, 7),
        "BIBLIOGRAFIA": html_bibliografia(ws_b),
    }
    injectar(html, seccions)
    print(f"  ✓  plm_tesi.html  →  {pg:.0f}% | tasques {task_done}/{task_total}")

def processar_elevador():
    xlsx = BASE / "elevador_plats" / "PLM_Elevador_Plats.xlsx"
    html = BASE / "elevador_plats" / "plm_elevador.html"
    if not xlsx.exists(): print("  ⚠  Excel no trobat"); return
    wb = load_workbook(xlsx, data_only=True)
    ws_f = wb["Fases"]; ws_t = wb["Tasques"]
    ws_r = wb["Riscos"]; ws_p = wb["Pressupost"]
    pg = progress_global(ws_f, 6)
    fases_done = sum(1 for r in files_dades(ws_f) if cel(ws_f,r,7) == "Completat")
    task_done  = sum(1 for r in files_dades(ws_t) if cel(ws_t,r,5) == "Completat")
    task_total = len(files_dades(ws_t))
    cost_est = 0
    for r in files_dades(ws_p):
        try:
            v = float(cel(ws_p, r, 6))
            if "TOTAL" not in str(cel(ws_p, r, 2)):
                cost_est += v
        except:
            pass
    seccions = {
        "PROGRESS_PCT": f'{pg:.0f}%',
        "PROGRESS_BAR": f'{pg:.0f}',
        "H_PROGRESS":   f'{pg:.0f}%',
        "H_FASES":      f'{fases_done}/{len(files_dades(ws_f))}',
        "H_TASQUES":    f'{task_done}/{task_total}',
        "H_COST":       f'{cost_est:.0f}€' if cost_est > 0 else "—",
        "FASES":        html_fases(ws_f, 2, 3, 6, 7, "#5C4A1A"),
        "TASQUES":      html_tasques(ws_t, 2, 3, 4, 5, 6, 7),
        "RISCOS":       html_riscos(ws_r, 2, 3, 4, 5, 7),
        "PRESSUPOST":   html_pressupost(ws_p, 2, 3, 6, 8),
    }
    injectar(html, seccions)
    print(f"  ✓  plm_elevador.html  →  {pg:.0f}% | tasques {task_done}/{task_total}")

def processar_frances():
    xlsx = BASE / "frances" / "PLM_Frances.xlsx"
    html = BASE / "frances" / "plm_frances.html"
    if not xlsx.exists(): print("  ⚠  Excel no trobat"); return
    wb = load_workbook(xlsx, data_only=True)
    ws_f = wb["Fases"]; ws_t = wb["Tasques"]
    ws_g = wb["Gramàtica"]; ws_v = wb["Vocabulari"]
    ws_r = wb["Registre"]
    pg = progress_global(ws_f, 6)
    task_done  = sum(1 for r in files_dades(ws_t) if cel(ws_t,r,5) == "Completat")
    task_total = len(files_dades(ws_t))
    vocab_apres  = sum(1 for r in files_dades(ws_v)
                       if str(cel(ws_v,r,6)) == "Après" and cel(ws_v,r,2))
    vocab_repas  = sum(1 for r in files_dades(ws_v)
                       if str(cel(ws_v,r,6)) == "En repàs" and cel(ws_v,r,2))
    vocab_pendent= sum(1 for r in files_dades(ws_v)
                       if str(cel(ws_v,r,6)) == "Pendent" and cel(ws_v,r,2))
    # Hores totals del registre
    hores_total = 0
    for r in files_dades(ws_r):
        try:
            hores_total += float(cel(ws_r, r, 4))
        except:
            pass
    seccions = {
        "PROGRESS_PCT":   f'{pg:.0f}%',
        "PROGRESS_BAR":   f'{pg:.0f}',
        "H_PROGRESS":     f'{pg:.0f}%',
        "H_HORES":        f'{hores_total:.1f}h',
        "H_VOCAB":        str(vocab_apres),
        "H_TASQUES":      f'{task_done}/{task_total}',
        "FASES":          html_fases(ws_f, 2, 3, 6, 7, "#1B3A5C"),
        "GRAMATICA":      html_gramatica(ws_g),
        "TASQUES":        html_tasques(ws_t, 2, 3, 4, 5, 6, 7),
        "VOCABULARI":     html_vocabulari(ws_v),
        "VOCAB_APRES":    str(vocab_apres),
        "VOCAB_REPAS":    str(vocab_repas),
        "VOCAB_PENDENT":  str(vocab_pendent),
        "REGISTRE":       html_registre(ws_r),
    }
    injectar(html, seccions)
    print(f"  ✓  plm_frances.html  →  {pg:.0f}% | vocab après: {vocab_apres} | hores: {hores_total:.1f}h")

def actualitzar_index():
    index = BASE / "index.html"
    if not index.exists(): return
    progres = {}
    for proj_id, xlsx_path, full_fases, c_pct in [
        ("web",  BASE/"web_personal"/"PLM_Web_Personal.xlsx",   "Fases",    6),
        ("tfm",  BASE/"tfm_quantica"/"PLM_TFM_Quantica.xlsx",   "Capítols", 6),
        ("elev", BASE/"elevador_plats"/"PLM_Elevador_Plats.xlsx","Fases",    6),
        ("fr",   BASE/"frances"/"PLM_Frances.xlsx",              "Fases",    6),
    ]:
        try:
            wb = load_workbook(xlsx_path, data_only=True)
            ws = wb[full_fases]
            progres[proj_id] = progress_global(ws, c_pct)
        except:
            progres[proj_id] = 0
    seccions = {
        "PCT_WEB":  f'{progres["web"]:.0f}%',
        "BAR_WEB":  f'{progres["web"]:.0f}',
        "PCT_TFM":  f'{progres["tfm"]:.0f}%',
        "BAR_TFM":  f'{progres["tfm"]:.0f}',
        "PCT_ELEV": f'{progres["elev"]:.0f}%',
        "BAR_ELEV": f'{progres["elev"]:.0f}',
        "PCT_FR":   f'{progres["fr"]:.0f}%',
        "BAR_FR":   f'{progres["fr"]:.0f}',
    }
    injectar(index, seccions)
    print(f"  ✓  index.html  →  web {progres['web']:.0f}% | tfm {progres['tfm']:.0f}% | elev {progres['elev']:.0f}% | fr {progres['fr']:.0f}%")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n=== Actualitzant portafoli ===\n")
    print("[Web Personal]");    processar_web()
    print("\n[TFM Quàntica]");  processar_tfm()
    print("\n[Elevador]");      processar_elevador()
    print("\n[Francès]");       processar_frances()
    print("\n[Pàgina principal]"); actualitzar_index()
    print("\n✓ Tot actualitzat. Fes git push per publicar.\n")

if __name__ == "__main__":
    main()
